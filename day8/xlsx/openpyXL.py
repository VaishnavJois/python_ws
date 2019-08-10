from openpyxl import Workbook
import openpyxl as pyxl
wb1 = Workbook()
sheet = wb1.active
data = [
    ('id','wname','year','status','company'),
    (1001,'Python',2019,1,'Heraizen'),
    (1002,'UI',2018,0,'Spaneos')
    ]

for row in data:
    sheet.append(row)
wb1.save('wsxl.xlsx')


wb2 = pyxl.load_workbook("wsxl.xlsx")
sheet2 = wb2.active
for row in sheet2.iter_rows():
    data = [c.value for c in row]
    print(data)

